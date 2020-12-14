# -*- coding: utf-8 -*-

#---------------------------------IMPORTANT!!!-------------------------------------
#
# Rename .xlsx file to "data.xlsx" and place it in the same directory with .py file.
#
# Cause .py file will find .xlsx file to read, I have to make sure that there is
# a .xlsx file in the same directory with .py file. So I rename the .xlsx file to
# "data.xlsx" and write code to check if it meets the condition.
#
# Idea: Using openpyxl to read and store data to a dictionary of category and
# value, and then calculate with these numbers.
#


import operator
import itertools
import os
import sys
import time
import pkg_resources
import subprocess
import platform

# Check if modules Openpyxl and matplotlib are installed or not, and automatically install if necessary
def install(pkg):
    if platform.system() == "Windows":
        subprocess.check_call([sys.executable,"-m","pip","install",pkg])
    elif platform.system() == "Linux":
    	subprocess.check_call(["pip3","install",pkg])
    else:
    	print("Sorry, I do not use MacOS so I do not know how to install a module in MacOS environment,")
    	print("so I suggest installing manually by running command on terminal.")
    	sys.exit()

for pkg in ["openpyxl","matplotlib"]:
	try:
		dist = pkg_resources.get_distribution(pkg)
		print("Module {} is installed, found version {}.".format(dist.key,dist.version))
	except pkg_resources.DistributionNotFound:
		print("{} is NOT installed, trying to install necessary packages...".format(pkg))
		install(pkg)
print("---------------\n")

import openpyxl
from matplotlib import pyplot as plt


# Check if .xlsx file is in the same directory with .py file
if not(os.path.isfile('data.xlsx')):
    print("Please check location of your .xlsx file, make sure it is named as 'data.xlsx' and in the same directory with .py file.")
    print("Current working directory: {}".format(os.getcwd()))
    path = os.getcwd()
    if platform.system() == "Windows":
        subprocess.check_call(["explorer",path])
    elif platform.system() == "Linux":
        subprocess.check_call(["xdg-open",path])
    elif platform.system() == "Darwin":
        subprocess.check_call(["open",path])
    sys.exit()
else:
    print("Found 'data.xlsx' file in {}, continue...\n".format(os.getcwd()))


# We create a dictionary to store data
d = {}
wb = openpyxl.load_workbook("data.xlsx")
sheet = wb["B14"]
for data in range(11,31):
    if (data == 17 or data == 19 or data == 20 or data == 21 or data == 22 or data == 23): continue
    category = sheet.cell(row = data, column =  2).value
    value = sheet.cell(row = data,column = 3).value
    d[category] = value
d[sheet.cell(row = 38, column = 2).value] = sheet.cell(row = 38,column = 3).value


# Sort dictionary X by value in descending order
def SortDictionaryByValue(x):
    sorted_x = dict(sorted(x.items(), key = operator.itemgetter(1),reverse = True))
    return sorted_x
sorted_d = SortDictionaryByValue(d)


# With integer k entered from keyboard, print k categories with highest
# and lowest values, and calculate the percentage of them in total

while True:
    k = input("Type integer k: ")
    try:
        k = int(k)
        if k <= 0:
            print("k must be a positive number.")
            continue
        break
    except ValueError:
        print("Integer, please.")


#Calculate the sum of all values
n = 0
for i in d:
    n = n + d[i]



# Print k categories with LOWEST values
print("{} categories with lowest values:\n".format(k))
b = itertools.islice(sorted_d.items(), len(sorted_d)-k, len(sorted_d))
for (key,value) in b: print(" -  ",key,"\t",value,"\t","{:.2f}%".format(float(value/n)*100))
print("\n")


# Print k categories with HIGHEST values
print("{} categories with highest values:\n".format(k))
b = itertools.islice(sorted_d.items(), 0, k)
for (key,value) in b: print(" -  ",key,"\t",value,"\t","{:.2f}%".format(float(value/n)*100))
print("\n")
time.sleep(2)


# List k categories (minimum) with total percentage of value in total is at least 80%
s = 0.0
count = 0
print("Categories of which total percentage of value in total is at least 80%:\n")
for i,data in enumerate(sorted_d,start = 1):
    s = s + sorted_d[data]/n
    if s<0.8:
        count = count + 1
        print(" {}. ".format(i),data)
    else:
        count = count + 1
        print(" {}. ".format(i),data)
        break
if count>1:
	print("\nThere are {} categories.".format(count))
else:
    print("\nThere is 1 category.")
print("\n")
time.sleep(2)


# Calculate the percentage of categories marked as "Include Foreign Element (IFE)" in total
print("There are 4 categories including foreign element:\n")
print(" -  Thu từ khu vực doanh nghiệp có vốn đầu tư nước ngoài;")
print(" -  Thu từ dầu thô;")
print(" -  Thu cân đối từ hoạt động xuất nhập khẩu;")
print(" -  Thu viện trợ\n")
print("and they account for {:.2f}% of total value.".format(float(d["Thu từ dầu thô"]/n + d["Thu cân đối từ hoạt động xuất nhập khẩu"]/n + d["Thu từ khu vực doanh nghiệp có vốn đầu tư nước ngoài"]/n + d["Thu viện trợ"]/n)*100))
print("\n")


# With category entered from keyboard, print its estimated value
while True:
    category = input("Enter category: ")
    print("")
    if category in d:
        print("The estimated value of '{}' in 2019: {} billion Vietnamese Dong.".format(category,d[category]))
        break
    else:
        print("Sorry, there are no categories that match to your input.")
print("\n")
time.sleep(2)


# Graph plotting
print("Graph plotting...")
print("Exporting to result.pdf...")

height = []
category = []
c = 0
left = [2,4,6,8,10,12,14,16,18,20,22,24,26,28,30]
for i in sorted_d:
    height.append(d[i]/1000)
    c += 1
    category.append("data{}".format(c))
plt.figure(figsize = (40,20))
plt.bar(left,height,color = ("#14ace3","#f09c3c"))
plt.xticks(left,category,rotation = 45,fontsize = 20)
plt.xlabel("Categories",fontsize = 30)
plt.ylabel("Estimated Value (thousand billion dong)",fontsize = 30)
plt.title("Estimated Revenue of Industries in 2019",fontsize = 30)
plt.savefig("result.pdf")
plt.show()
print("The graph is saved and exported to result.pdf in {}.".format(os.getcwd()))

# --------------------------------BEGIN OF EXAMPLE OUTPUT-------------------------------
#
# Module openpyxl is installed, found version 3.0.3.
# Module matplotlib is installed, found version 3.2.1.
# ---------------
#
# Found 'data.xlsx' file in /*FILEPATH*/, continue...
#
# Type integer k: 4
# 4 categories with lowest values:
#
#  -   Thu khác ngân sách 	 24563 	 1.74%
#  -   Thu tiền cấp quyền khai thác khoáng sản, tài nguyên nước 	 4069 	 0.29%
#  -   Thu viện trợ 	 4000 	 0.28%
#  -   Thu từ quỹ đất công ích và thu hoa lợi công sản khác 	 927 	 0.07%
#
#
# 4 categories with highest values:
#
#  -   Thu từ khu vực kinh tế ngoài quốc doanh 	 241529.5 	 17.11%
#  -   Thu từ khu vực doanh nghiệp có vốn đầu tư nước ngoài 	 213733.5 	 15.14%
#  -   Thu cân đối từ hoạt động xuất nhập khẩu 	 189200 	 13.41%
#  -   Thu từ khu vực doanh nghiệp nhà nước 	 177709 	 12.59%
#
#
# Categories of which total percentage of value in total is at least 80%:
#
#  1.  Thu từ khu vực kinh tế ngoài quốc doanh
#  2.  Thu từ khu vực doanh nghiệp có vốn đầu tư nước ngoài
#  3.  Thu cân đối từ hoạt động xuất nhập khẩu
#  4.  Thu từ khu vực doanh nghiệp nhà nước
#  5.  Các khoản thu về nhà, đất
#  6.  Thuế thu nhập cá nhân
#  7.  Thu hồi vốn, thu cổ tức, lợi nhuận, lợi nhuận sau thuế, chênh lệch thu, chi của Ngân hàng Nhà nước
#
# There are 7 categories.
#
#
# There are 4 categories including foreign element:
#
#  -  Thu từ khu vực doanh nghiệp có vốn đầu tư nước ngoài;
#  -  Thu từ dầu thô;
#  -  Thu cân đối từ hoạt động xuất nhập khẩu;
#  -  Thu viện trợ
#
# and they account for 31.99% of total value.
#
#
# Enter category: Thu viện trợ
#
# The estimated value of 'Thu viện trợ' in 2019: 4000 billion Vietnamese Dong.
#
#--------------------------------END OF EXAMPLE OUTPUT----------------------------------
