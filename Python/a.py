import operator
import itertools
import os
import sys
import time
import pkg_resources
import subprocess
import platform
import openpyxl

# We create a dictionary to store data
wb = openpyxl.load_workbook("Sinno_recrutment.xlsx")
sheet = wb["SS"]
for data in range(1,21):
    value = sheet.cell(row = data,column = 1).value
    j=1
    i=0
    temp=""
    while i<len(value)-3:
        if value[i]=="'" and value[i+1]=="," and value[i+2]=="'":
            sheet.cell(row = data,column = j).value = temp
            j=j+1
            i=i+3
            temp=""
        else:
            temp=temp+value[i]
            i=i+1

sheet = wb["CCC"]
for data in range(1,113):
    value = sheet.cell(row = data,column = 4).value
    j=1
    i=0
    temp=""
    while i<len(value)-3:
        if value[i]=="'" and value[i+1]=="," and value[i+2]=="'":
            sheet.cell(row = data,column = j).value = temp
            j=j+1
            i=i+3
            temp=""
        else:
            temp=temp+value[i]
            i=i+1

sheet = wb["A"]
for data in range(1,96):
    value = sheet.cell(row = data,column = 1).value
    j=1
    i=0
    temp=""
    while i<len(value)-3:
        if value[i]=="'" and value[i+1]=="," and value[i+2]=="'":
            sheet.cell(row = data,column = j).value = temp
            j=j+1
            i=i+3
            temp=""
        else:
            temp=temp+value[i]
            i=i+1

sheet = wb["C"]
for data in range(1,34):
    value = sheet.cell(row = data,column = 1).value
    j=1
    i=0
    temp=""
    while i<len(value)-3:
        if value[i]=="'" and value[i+1]=="," and value[i+2]=="'":
            sheet.cell(row = data,column = j).value = temp
            j=j+1
            i=i+3
            temp=""
        else:
            temp=temp+value[i]
            i=i+1

sheet = wb["B"]
for data in range(1,33):
    value = sheet.cell(row = data,column = 1).value
    j=1
    i=0
    temp=""
    while i<len(value)-3:
        if value[i]=="'" and value[i+1]=="," and value[i+2]=="'":
            sheet.cell(row = data,column = j).value = temp
            j=j+1
            i=i+3
            temp=""
        else:
            temp=temp+value[i]
            i=i+1
sheet = wb["All"]
for data in range(1,295):
    value = sheet.cell(row = data,column = 1).value
    j=1
    i=0
    temp=""
    while i<len(value)-3:
        if value[i]=="'" and value[i+1]=="," and value[i+2]=="'":
            sheet.cell(row = data,column = j).value = temp
            j=j+1
            i=i+3
            temp=""
        else:
            temp=temp+value[i]
            i=i+1

wb.save("a.xlsx")